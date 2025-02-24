import openpyxl

def add_instruction_column(file_path, new_file_path):
    # 打开Excel文件
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # 插入新列
    ws.insert_cols(1)
    
    # 设置标题
    ws.cell(row=1, column=1, value='Instruction')
    
    # 填充数据并删除特定行
    rows_to_delete = []
    for row in range(2, ws.max_row + 1):
        google_translation_value = ws.cell(row=row, column=4).value
        col6_value = ws.cell(row=row, column=6).value
        col7_value = ws.cell(row=row, column=7).value
        
        if google_translation_value == ": : : : : : : : : : : : : :" or (col6_value is not None and col7_value is not None and col6_value < 60 and col7_value < 60):
            rows_to_delete.append(row)
        else:
            if ws.cell(row=row, column=2).value is not None:
                ws.cell(row=row, column=1, value='''你是一位专业的金融领域中英互译专家。请将以下中文金融文本翻译成英文。
    要求：
    1. 准确使用金融专业术语，保持专业性
    2. 严格遵循金融行业通用表达方式
    3. 保持原文的金融概念完整性
    4. 确保数字、百分比等关键信息的准确传达
    5. 保持报表和数据分析相关术语的统一性
    ''')
    
    # 删除标记的行
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)
    
    # 保存为新文件
    wb.save(new_file_path)

# 示例调用
add_instruction_column('./out/finance_1_evaluation_results_.xlsx', './out/金融类数据翻译结果.xlsx')
